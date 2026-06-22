import csv

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import ARQUIVO_CSV, ARQUIVO_EXCEL


def exportar_clientes_csv(clientes):
    with open(ARQUIVO_CSV, "w", encoding="utf-8", newline="") as arquivo:
        campos = ["nome", "email", "telefone"]

        escritor = csv.DictWriter(arquivo, fieldnames=campos)

        escritor.writeheader()
        escritor.writerows(clientes)

    print(f"Arquivo CSV de clientes gerado com sucesso: {ARQUIVO_CSV}")


def exportar_clientes_excel(clientes):
    planilha_excel = Workbook()

    pagina_clientes = planilha_excel.active
    pagina_clientes.title = "Clientes"

    pagina_clientes.append(["Nome", "E-mail", "Telefone"])

    for cliente in clientes:
        pagina_clientes.append([
            cliente["nome"],
            cliente["email"],
            str(cliente["telefone"])
        ])

    for celula in pagina_clientes[1]:
        celula.font = Font(bold=True)
        celula.alignment = Alignment(horizontal="center", vertical="center")

    for linha in pagina_clientes.iter_rows(
        min_row=2,
        max_row=pagina_clientes.max_row,
        min_col=1,
        max_col=3
    ):
        for celula in linha:
            celula.alignment = Alignment(vertical="center")

    for celula in pagina_clientes["C"]:
        celula.number_format = "@"

    pagina_clientes.column_dimensions["A"].width = 25
    pagina_clientes.column_dimensions["B"].width = 35
    pagina_clientes.column_dimensions["C"].width = 18

    for numero_linha in range(1, pagina_clientes.max_row + 1):
        pagina_clientes.row_dimensions[numero_linha].height = 22

    pagina_clientes.freeze_panes = "A2"

    intervalo_tabela = f"A1:C{pagina_clientes.max_row}"

    tabela_clientes = Table(
        displayName="TabelaClientes",
        ref=intervalo_tabela
    )

    estilo_tabela = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    tabela_clientes.tableStyleInfo = estilo_tabela

    pagina_clientes.add_table(tabela_clientes)

    planilha_excel.save(ARQUIVO_EXCEL)

    print(f"Planilha Excel de clientes gerada com sucesso: {ARQUIVO_EXCEL}")